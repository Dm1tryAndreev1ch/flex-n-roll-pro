"""
Экспорт аналитических данных в Excel с форматированием.
"""

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, numbers,
)
from openpyxl.utils.dataframe import dataframe_to_rows

from config.settings import REPORTS_OUTPUT_DIR
from src.utils.logger import get_logger

logger = get_logger("reports.excel_export")

# ── Стили ──

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F4E79")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11, color="2E75B6")

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

RISK_FILLS = {
    "HIGH": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
    "MEDIUM": PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid"),
    "LOW": PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid"),
}

SEGMENT_FILLS = {
    "Champions": PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid"),
    "Loyal Customers": PatternFill(start_color="45B7D1", end_color="45B7D1", fill_type="solid"),
    "Promising": PatternFill(start_color="96CEB4", end_color="96CEB4", fill_type="solid"),
    "New Customers": PatternFill(start_color="DDA0DD", end_color="DDA0DD", fill_type="solid"),
    "At Risk": PatternFill(start_color="FFEAA7", end_color="FFEAA7", fill_type="solid"),
    "Lost": PatternFill(start_color="FF7675", end_color="FF7675", fill_type="solid"),
}

MONEY_FORMAT = '#,##0.00 "BYN"'
PERCENT_FORMAT = "0.0%"
NUMBER_FORMAT = "#,##0"


class ExcelExporter:
    """
    Экспорт аналитических данных в Excel (.xlsx).

    Листы:
    1. «Обзор» — KPI дашборд
    2. «RFM-сегменты» — таблица RFM-сегментации
    3. «Риск оттока» — клиенты с HIGH риском
    4. «Прогноз» — таблица и линейный график
    """

    def export(
        self,
        forecast_df: pd.DataFrame | None = None,
        rfm_df: pd.DataFrame | None = None,
        churn_df: pd.DataFrame | None = None,
        deals_df: pd.DataFrame | None = None,
        output_filename: str | None = None,
    ) -> Path:
        """
        Экспортировать данные в Excel.

        Args:
            forecast_df: Прогноз выручки.
            rfm_df: RFM-сегментация.
            churn_df: Предсказание оттока.
            deals_df: Сделки (для листа «Обзор»).
            output_filename: Имя файла (без расширения).

        Returns:
            Путь к созданному файлу.
        """
        wb = Workbook()
        wb.remove(wb.active)  # Удаляем дефолтный лист

        # 1. Лист «Обзор»
        if deals_df is not None:
            self._create_overview_sheet(wb, deals_df, rfm_df, churn_df, forecast_df)

        # 2. Лист «RFM-сегменты»
        if rfm_df is not None and not rfm_df.empty:
            self._create_rfm_sheet(wb, rfm_df)

        # 3. Лист «Риск оттока»
        if churn_df is not None and not churn_df.empty:
            self._create_churn_sheet(wb, churn_df)

        # 4. Лист «Прогноз»
        if forecast_df is not None and not forecast_df.empty:
            self._create_forecast_sheet(wb, forecast_df)

        # Если ни один лист не создан
        if not wb.sheetnames:
            ws = wb.create_sheet("Нет данных")
            ws["A1"] = "Нет данных для экспорта"

        # Сохранение
        output_dir = Path(REPORTS_OUTPUT_DIR)
        output_dir.mkdir(parents=True, exist_ok=True)

        if not output_filename:
            output_filename = f"fnr_analytics_{pd.Timestamp.now().strftime('%Y-%m-%d_%H%M')}"
        filepath = output_dir / f"{output_filename}.xlsx"

        wb.save(filepath)
        logger.info("Excel-отчёт сохранён: %s", filepath)
        return filepath

    # ──────────────────────────────────
    #  Лист «Обзор»
    # ──────────────────────────────────

    def _create_overview_sheet(
        self,
        wb: Workbook,
        deals_df: pd.DataFrame,
        rfm_df: pd.DataFrame | None,
        churn_df: pd.DataFrame | None,
        forecast_df: pd.DataFrame | None,
    ) -> None:
        """Создать лист «Обзор» с KPI-дашбордом."""
        ws = wb.create_sheet("Обзор")

        # Заголовок
        ws.merge_cells("A1:F1")
        cell = ws["A1"]
        cell.value = "FLEX-N-ROLL PRO — Аналитический дашборд"
        cell.font = TITLE_FONT
        cell.alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:F2")
        ws["A2"].value = f"Дата: {pd.Timestamp.now().strftime('%d.%m.%Y %H:%M')}"
        ws["A2"].font = SUBTITLE_FONT
        ws["A2"].alignment = Alignment(horizontal="center")

        # KPI блок
        row = 4
        ws.cell(row=row, column=1, value="Ключевые показатели").font = SUBTITLE_FONT
        row += 1

        won_deals = deals_df[deals_df["is_won"]] if "is_won" in deals_df.columns else pd.DataFrame()
        lost_deals = deals_df[deals_df["is_lost"]] if "is_lost" in deals_df.columns else pd.DataFrame()
        active_deals = deals_df[deals_df["is_active"]] if "is_active" in deals_df.columns else pd.DataFrame()

        kpis = [
            ("Всего сделок", len(deals_df)),
            ("Выигранных", len(won_deals)),
            ("Проигранных", len(lost_deals)),
            ("В работе", len(active_deals)),
            ("Общая выручка (BYN)", won_deals["OPPORTUNITY"].sum() if not won_deals.empty else 0),
            ("Средний чек (BYN)", won_deals["OPPORTUNITY"].mean() if not won_deals.empty else 0),
            ("Win Rate", f"{len(won_deals)/(len(won_deals)+len(lost_deals))*100:.1f}%" if (len(won_deals) + len(lost_deals)) > 0 else "—"),
        ]

        if churn_df is not None and not churn_df.empty:
            high_risk = len(churn_df[churn_df["risk_level"] == "HIGH"])
            kpis.append(("Клиентов HIGH-риска оттока", high_risk))

        if rfm_df is not None and not rfm_df.empty:
            champions = len(rfm_df[rfm_df["segment"] == "Champions"])
            kpis.append(("Champions (RFM)", champions))

        for kpi_name, kpi_value in kpis:
            ws.cell(row=row, column=1, value=kpi_name).font = Font(name="Calibri", bold=True)
            ws.cell(row=row, column=1).border = THIN_BORDER
            cell = ws.cell(row=row, column=2, value=kpi_value)
            cell.border = THIN_BORDER
            if isinstance(kpi_value, float):
                cell.number_format = MONEY_FORMAT
            row += 1

        # Распределение по категориям
        row += 1
        if "product_category" in deals_df.columns and not won_deals.empty:
            ws.cell(row=row, column=1, value="Выручка по категориям продукции").font = SUBTITLE_FONT
            row += 1

            cat_data = won_deals.groupby("product_category")["OPPORTUNITY"].sum().sort_values(ascending=False)
            headers = ["Категория", "Выручка (BYN)", "Доля (%)"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = HEADER_ALIGNMENT
                cell.border = THIN_BORDER
            row += 1

            total = cat_data.sum()
            for cat, rev in cat_data.items():
                ws.cell(row=row, column=1, value=cat).border = THIN_BORDER
                c = ws.cell(row=row, column=2, value=rev)
                c.number_format = MONEY_FORMAT
                c.border = THIN_BORDER
                c = ws.cell(row=row, column=3, value=rev / total if total > 0 else 0)
                c.number_format = PERCENT_FORMAT
                c.border = THIN_BORDER
                row += 1

        # Автоширина колонок
        for col in range(1, 7):
            ws.column_dimensions[chr(64 + col)].width = 22

    # ──────────────────────────────────
    #  Лист «RFM-сегменты»
    # ──────────────────────────────────

    def _create_rfm_sheet(self, wb: Workbook, rfm_df: pd.DataFrame) -> None:
        """Создать лист RFM-сегментации."""
        ws = wb.create_sheet("RFM-сегменты")

        # Заголовок
        ws.merge_cells("A1:K1")
        ws["A1"].value = "RFM-сегментация клиентов"
        ws["A1"].font = TITLE_FONT
        ws["A1"].alignment = Alignment(horizontal="center")

        # Сводка по сегментам
        row = 3
        ws.cell(row=row, column=1, value="Сводка по сегментам").font = SUBTITLE_FONT
        row += 1

        seg_headers = ["Сегмент", "Количество", "Доля", "Выручка (BYN)"]
        for col_idx, h in enumerate(seg_headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
        row += 1

        total_count = len(rfm_df)
        for segment in rfm_df["segment"].unique():
            seg_data = rfm_df[rfm_df["segment"] == segment]
            c1 = ws.cell(row=row, column=1, value=segment)
            c1.border = THIN_BORDER
            if segment in SEGMENT_FILLS:
                c1.fill = SEGMENT_FILLS[segment]

            ws.cell(row=row, column=2, value=len(seg_data)).border = THIN_BORDER
            c3 = ws.cell(row=row, column=3, value=len(seg_data)/total_count if total_count > 0 else 0)
            c3.number_format = PERCENT_FORMAT
            c3.border = THIN_BORDER
            c4 = ws.cell(row=row, column=4, value=seg_data["monetary"].sum())
            c4.number_format = MONEY_FORMAT
            c4.border = THIN_BORDER
            row += 1

        # Детальная таблица
        row += 1
        ws.cell(row=row, column=1, value="Детальная таблица").font = SUBTITLE_FONT
        row += 1

        display_cols = [
            "company_name", "segment", "recency", "frequency", "monetary",
            "r_score", "f_score", "m_score", "rfm_score",
        ]
        col_labels = [
            "Компания", "Сегмент", "Recency (дн.)", "Frequency", "Monetary (BYN)",
            "R-скор", "F-скор", "M-скор", "RFM",
        ]

        for col_idx, label in enumerate(col_labels, 1):
            cell = ws.cell(row=row, column=col_idx, value=label)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER
        row += 1

        for _, data_row in rfm_df.iterrows():
            for col_idx, col in enumerate(display_cols, 1):
                val = data_row.get(col, "")
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.border = THIN_BORDER
                if col == "monetary":
                    cell.number_format = MONEY_FORMAT
                if col == "segment" and val in SEGMENT_FILLS:
                    cell.fill = SEGMENT_FILLS[val]
            row += 1

        # Автоширина
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 18
        for col in "CDEFGHI":
            ws.column_dimensions[col].width = 15

    # ──────────────────────────────────
    #  Лист «Риск оттока»
    # ──────────────────────────────────

    def _create_churn_sheet(self, wb: Workbook, churn_df: pd.DataFrame) -> None:
        """Создать лист с рисками оттока."""
        ws = wb.create_sheet("Риск оттока")

        # Заголовок
        ws.merge_cells("A1:I1")
        ws["A1"].value = "Предсказание риска оттока клиентов"
        ws["A1"].font = TITLE_FONT
        ws["A1"].alignment = Alignment(horizontal="center")

        # Сводка
        row = 3
        ws.cell(row=row, column=1, value="Сводка по уровням риска").font = SUBTITLE_FONT
        row += 1

        for risk_level in ["HIGH", "MEDIUM", "LOW"]:
            count = len(churn_df[churn_df["risk_level"] == risk_level])
            cell = ws.cell(row=row, column=1, value=f"{risk_level}: {count} компаний")
            cell.font = Font(name="Calibri", bold=True, size=11)
            if risk_level in RISK_FILLS:
                cell.fill = RISK_FILLS[risk_level]
            cell.border = THIN_BORDER
            row += 1

        # Детальная таблица
        row += 1
        ws.cell(row=row, column=1, value="Детали (все клиенты)").font = SUBTITLE_FONT
        row += 1

        display_cols = [
            "company_name", "churn_probability", "risk_level",
            "last_order_date", "days_since_last_order",
            "total_orders", "avg_order_value", "recommended_action",
        ]
        col_labels = [
            "Компания", "Вер-ть оттока", "Риск",
            "Последний заказ", "Дней без заказов",
            "Всего заказов", "Средний чек", "Рекомендация",
        ]

        for col_idx, label in enumerate(col_labels, 1):
            cell = ws.cell(row=row, column=col_idx, value=label)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER
        row += 1

        for _, data_row in churn_df.iterrows():
            for col_idx, col in enumerate(display_cols, 1):
                val = data_row.get(col, "")
                if isinstance(val, pd.Timestamp):
                    val = val.strftime("%d.%m.%Y")
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.border = THIN_BORDER

                if col == "churn_probability":
                    cell.number_format = "0.0%"
                elif col == "avg_order_value":
                    cell.number_format = MONEY_FORMAT
                elif col == "risk_level" and val in RISK_FILLS:
                    cell.fill = RISK_FILLS[val]
                    cell.font = Font(name="Calibri", bold=True)
            row += 1

        # Автоширина
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 16
        ws.column_dimensions["F"].width = 14
        ws.column_dimensions["G"].width = 16
        ws.column_dimensions["H"].width = 50

    # ──────────────────────────────────
    #  Лист «Прогноз»
    # ──────────────────────────────────

    def _create_forecast_sheet(self, wb: Workbook, forecast_df: pd.DataFrame) -> None:
        """Создать лист с прогнозом и графиком."""
        ws = wb.create_sheet("Прогноз")

        # Заголовок
        ws.merge_cells("A1:F1")
        ws["A1"].value = "Прогноз выручки"
        ws["A1"].font = TITLE_FONT
        ws["A1"].alignment = Alignment(horizontal="center")

        # Таблица данных
        row = 3
        headers = ["Месяц", "Факт (BYN)", "Прогноз (BYN)", "Нижняя граница CI", "Верхняя граница CI", "Тип"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER
        row += 1

        data_start_row = row
        forecast_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")

        for _, data_row in forecast_df.iterrows():
            ws.cell(row=row, column=1, value=data_row["month"]).border = THIN_BORDER

            actual = data_row.get("actual")
            c2 = ws.cell(row=row, column=2, value=actual if pd.notna(actual) else None)
            c2.number_format = MONEY_FORMAT
            c2.border = THIN_BORDER

            c3 = ws.cell(row=row, column=3, value=data_row["forecast"])
            c3.number_format = MONEY_FORMAT
            c3.border = THIN_BORDER

            c4 = ws.cell(row=row, column=4, value=data_row["lower_ci"])
            c4.number_format = MONEY_FORMAT
            c4.border = THIN_BORDER

            c5 = ws.cell(row=row, column=5, value=data_row["upper_ci"])
            c5.number_format = MONEY_FORMAT
            c5.border = THIN_BORDER

            type_label = "Прогноз" if data_row.get("is_forecast") else "Факт"
            ws.cell(row=row, column=6, value=type_label).border = THIN_BORDER

            if data_row.get("is_forecast"):
                for col in range(1, 7):
                    ws.cell(row=row, column=col).fill = forecast_fill

            row += 1

        data_end_row = row - 1

        # Линейный график
        if data_end_row >= data_start_row:
            chart = LineChart()
            chart.title = "Прогноз выручки FLEX-N-ROLL PRO"
            chart.y_axis.title = "Выручка (BYN)"
            chart.x_axis.title = "Месяц"
            chart.width = 28
            chart.height = 14
            chart.style = 10

            cats = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_end_row)

            # Линия факта
            actual_data = Reference(ws, min_col=2, min_row=data_start_row - 1, max_row=data_end_row)
            chart.add_data(actual_data, titles_from_data=True)
            chart.set_categories(cats)

            # Линия прогноза
            forecast_data = Reference(ws, min_col=3, min_row=data_start_row - 1, max_row=data_end_row)
            chart.add_data(forecast_data, titles_from_data=True)

            # Доверительные интервалы
            lower_data = Reference(ws, min_col=4, min_row=data_start_row - 1, max_row=data_end_row)
            chart.add_data(lower_data, titles_from_data=True)

            upper_data = Reference(ws, min_col=5, min_row=data_start_row - 1, max_row=data_end_row)
            chart.add_data(upper_data, titles_from_data=True)

            # Стили линий
            if len(chart.series) > 0:
                chart.series[0].graphicalProperties.line.width = 25000  # Факт — жирная
            if len(chart.series) > 1:
                chart.series[1].graphicalProperties.line.dashStyle = "dash"  # Прогноз — пунктир
            if len(chart.series) > 2:
                chart.series[2].graphicalProperties.line.dashStyle = "dot"
                chart.series[2].graphicalProperties.line.solidFill = "AAAAAA"
            if len(chart.series) > 3:
                chart.series[3].graphicalProperties.line.dashStyle = "dot"
                chart.series[3].graphicalProperties.line.solidFill = "AAAAAA"

            ws.add_chart(chart, f"A{row + 2}")

        # Автоширина
        for col_letter in "ABCDEF":
            ws.column_dimensions[col_letter].width = 20
